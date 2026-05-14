import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Generador de Pedidos MOTSUR", layout="wide")

# Diccionario de costos por pulgadas
COSTOS_REPUESTOS = {
    '43': 130,
    '50': 150,
    '55': 170,
    '65': 240,
    '75': 350,
    '85': 700
}

def extraer_codigo_final(nombre_producto):
    if pd.isna(nombre_producto): return "S/N"
    nombre_str = str(nombre_producto).upper().replace(" ", "").replace("-", "")
    nombre_str = nombre_str.replace("4K", "")
    match_numero = re.search(r'\d', nombre_str)
    if match_numero:
        inicio = match_numero.start()
        bloque = nombre_str[inicio:inicio+5]
        return f"PL{bloque}"
    return "SINMODELO"

def calcular_costo_item(modelo):
    """Detecta las pulgadas en el modelo y retorna el costo."""
    modelo_str = str(modelo)
    for pulgada, costo in COSTOS_REPUESTOS.items():
        if pulgada in modelo_str:
            return costo
    return 0  # Si no detecta pulgadas conocidas

st.title("📊 Generador de Pedidos - Control de Presupuesto ($2600)")

uploaded_file = st.file_uploader("Sube el archivo Excel de control", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        df_master = pd.read_excel(uploaded_file)
        df_master.columns = [str(c).strip() for c in df_master.columns]
        columnas_reales = list(df_master.columns)
        
        def detectar(keywords):
            for i, col in enumerate(columnas_reales):
                if any(k.upper() in col.upper() for k in keywords):
                    return i
            return 0

        st.info("Configuración de Columnas:")
        c1, c2 = st.columns(2)
        with c1:
            sel_orden = st.selectbox("1. ORDEN:", columnas_reales, index=detectar(['ORDEN', '#']))
            sel_serie = st.selectbox("2. SERIE:", columnas_reales, index=detectar(['SERIE']))
            sel_modelo = st.selectbox("3. MODELO:", columnas_reales, index=detectar(['PRODUCTO', 'MODELO']))
        with c2:
            sel_procedencia = st.selectbox("4. TALLER DE PROCEDENCIA:", columnas_reales, index=detectar(['PROCEDENCIA']))
            sel_taller = st.selectbox("5. TALLER:", columnas_reales, index=detectar(['TECNICO', 'TALLER']))
            sel_repuesto = st.selectbox("6. REPUESTO:", columnas_reales, index=detectar(['REPUESTO']))

        st.divider()
        input_ordenes = st.text_area("Pegue las órdenes aquí (una por línea):", height=150)

        if st.button("🚀 Procesar con Límite de Presupuesto", type="primary"):
            if input_ordenes:
                lista_busqueda = [o.strip() for o in input_ordenes.split('\n') if o.strip()]
                df_master[sel_orden] = df_master[sel_orden].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                
                # Filtrar y mantener el orden en que se pegaron las órdenes
                df_filtrado = df_master[df_master[sel_orden].isin(lista_busqueda)].copy()
                
                # Lógica de presupuesto
                ordenes_aceptadas = []
                costo_acumulado = 0
                limite_maximo = 2600
                ordenes_excluidas = 0

                for _, fila in df_filtrado.iterrows():
                    costo_actual = calcular_costo_item(fila[sel_modelo])
                    if (costo_acumulado + costo_actual) <= limite_maximo:
                        costo_acumulado += costo_actual
                        ordenes_aceptadas.append(fila)
                    else:
                        ordenes_excluidas += 1

                if ordenes_aceptadas:
                    df_res = pd.DataFrame(ordenes_aceptadas)
                    df_res['CODIGO_PL'] = df_res[sel_modelo].apply(extraer_codigo_final)
                    
                    df_final = pd.DataFrame({
                        'ORDEN': df_res[sel_orden],
                        'SERIE': df_res[sel_serie],
                        'MODELO': df_res[sel_modelo],
                        'TALLER DE PROCEDENCIA': df_res[sel_procedencia],
                        'TALLER': df_res[sel_taller],
                        'REPUESTO': df_res[sel_repuesto],
                        'CODIGO': df_res['CODIGO_PL']
                    }).fillna("REVISAR")

                    # Mostrar resumen de presupuesto
                    st.success(f"✅ Pedido generado. Presupuesto utilizado: ${costo_acumulado} / ${limite_maximo}")
                    if ordenes_excluidas > 0:
                        st.warning(f"⚠️ Se excluyeron {ordenes_excluidas} órdenes por superar el límite de $2600.")

                    st.dataframe(df_final, use_container_width=True)

                    # --- EXCEL PROFESIONAL ---
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_final.to_excel(writer, index=False, sheet_name='PEDIDO')
                        ws = writer.sheets['PEDIDO']
                        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        center_align = Alignment(horizontal="center", vertical="center")
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                        for col_num, column in enumerate(df_final.columns, 1):
                            cell = ws.cell(row=1, column=col_num)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_align
                            cell.border = thin_border
                            max_len = max(df_final[column].astype(str).map(len).max(), len(column)) + 5
                            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max_len

                        for row in ws.iter_rows(min_row=2, max_row=len(df_final)+1, max_col=7):
                            for cell in row:
                                cell.border = thin_border

                    data_excel = output.getvalue()
                    fecha_str = datetime.now().strftime('%d%m%Y')
                    nombre_archivo = f"PEDIDO BDG 64 MOTSUR TVS {fecha_str}.xlsx"

                    st.download_button(
                        label="📥 Descargar Excel MOTSUR TVS",
                        data=data_excel,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("No se pudieron procesar órdenes dentro del presupuesto.")
    except Exception as e:
        st.error(f"Error técnico: {e}")
